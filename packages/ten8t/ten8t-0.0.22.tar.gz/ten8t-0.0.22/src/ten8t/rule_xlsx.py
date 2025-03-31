"""
Module containing baseline rules for verifying the contents of Excel files.

This module is designed to provide a foundation for implementing rules that can analyze the
contents of Excel (.xlsx) files to ensure they adhere to expected formats and standards.
Methods or classes in this module can be extended or combined for more specific use cases.
"""

import openpyxl
import pandas as pd

from .render import BM
from .ten8t_exception import Ten8tException
from .ten8t_result import TR
from .ten8t_util import StrOrNone, str_to_bool

SHEET1 = "Sheet1"
AUTO = "auto"
DESC_DEFAULT = ""
START_ROW_DEFAULT = "1"
VAL_COL_DEFAULT = "B"
DESCRIPTION_COLUMN = "A"


def _column_to_number(column: str) -> int:
    """
    Converts an Excel-like column string into its corresponding column number.

    This function takes a string representing an Excel column label (e.g., 'A', 'Z',
    'AA') and converts it into the corresponding numeric representation. The
    computation is based on treating the column string as a base-26 number system,
    where 'A' corresponds to 1, 'B' corresponds to 2, and so on up to 'Z'. For
    multi-character strings, their positions are calculated cumulatively based on
    their place values.

    Args:
        column (str): The Excel-like column label to convert. Must consist of
                      uppercase or lowercase alphabetic characters.

    Returns:
        int: The numeric representation of the column label.
    """
    number = 0
    for i, char in enumerate(reversed(column)):
        number += (ord(char.upper()) - 64) * (26 ** i)
    return number


def _get_sheet(wb, sheet_name=None, first_if_missing=True):
    """Ensure a valid sheet is selected based on sheet_name parameter"""
    if sheet_name is None:
        return wb["Sheet1"]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # the sheet name is missing and the first_if_missing is set true then return that
    if len(wb.sheetnames) >= 1 and first_if_missing:
        return wb[wb.sheetnames[0]]

    raise Ten8tException('A sheet name was not specified and sheet1 could not be found.')


def _ensure_row_params(row_end, row_start: int):
    """
    Validates and processes row parameters, ensuring correctness and
    consistency. Adjusts the `row_end` value when it is set to AUTO and
    confirms the validity of provided values based on constraints.

    Args:
        row_end: The ending row, which can be an integer, a string that represents
            an integer, the special string value `AUTO`, or None.
        row_start (int): The starting row, specified as an integer, which must not
            be larger than the provided or derived `row_end`.

    Returns:
        tuple: A tuple containing the adjusted `row_start`, `row_end`, and a
        boolean flag `auto`. The `auto` flag indicates if the `row_end` was
        automatically adjusted to the predefined default value.

    Raises:
        Ten8tException: If `row_end` is a string representing an integer value less
            than `row_start`, or if `row_end` is invalid (neither an integer, a string
            representing a valid integer, nor the special value AUTO).
    """
    auto = False

    if row_end is None:
        return row_start, row_start, auto

    if isinstance(row_end, str):
        if row_end.isdigit() and int(row_end) < row_start:
            raise Ten8tException(
                f'Value for end row must be larger than start row {row_start=} {row_end=}')
        if row_end.lower() == AUTO:
            auto = True
            row_end = 1000
    try:
        row_end = int(row_end)
    except ValueError as vex:
        raise Ten8tException("row_end was not a valid integer value") from vex
    return row_start, row_end, auto


def rule_xlsx_a1_pass_fail(wb: openpyxl.workbook.Workbook,
                           sheet_name: StrOrNone = None,
                           desc_col='A',
                           val_col='B',
                           row_start='1',
                           row_end=None,
                           first_if_missing=True):
    """
    Processes an Excel sheet to evaluate rows and yield pass or fail results based on input values.

    The function reads through the rows of a specified Excel sheet, validates the presence of boolean-like values in a
    specified column, and returns pass or fail outputs for each row. Optionally, a description column can be specified
    to include textual descriptions in the result messages.

    Args:
        wb (openpyxl.workbook.Workbook): The workbook object containing the sheet to process.
        sheet_name (StrOrNone, optional): The name of the sheet to process. If None, the active sheet is used.
        desc_col (str, optional): The column identifier (e.g., 'A') for the description. Defaults to 'A'.
        val_col (str, optional): The column identifier (e.g., 'B') for the boolean-like values. Defaults to 'B'.
        row_start (str, optional): The starting row number as a string. Defaults to '1'.
        row_end (str or None, optional): The ending row number as a string or None to infer the range automatically.
        first_if_missing(bool): If no sheet is found then use the first one

    Yields:
        TR: An object indicating pass (True) or fail (False) status for each row and an accompanying message.

    Raises:
        Ten8tException: If a non-boolean-like value is encountered in the value column.
    """
    sheet = _get_sheet(wb, sheet_name, first_if_missing=first_if_missing)

    # Handle Nones.  Presumably this should not be required
    row_start = row_start or '1'
    val_col = val_col or 'B'
    row_start, row_end, auto = _ensure_row_params(row_end, int(row_start))
    val_col = _column_to_number(val_col)

    if desc_col is not None:
        desc_col = _column_to_number(desc_col)

    for row in range(row_start, row_end + 1):

        value = sheet.cell(row=row, column=val_col).value
        if value is None and auto:
            break

        if value is None:
            raise Ten8tException(f'Expected boolean value in row {row}')

        if desc_col is not None:
            desc = sheet.cell(row=row, column=desc_col).value
        else:
            # It is possible not to have a description column
            desc = ""

        if str_to_bool(value):
            yield TR(status=True, msg=f"{BM.expected(desc)}-Passed")
        else:
            yield TR(status=False, msg=f"{BM.expected(desc)}-Failed")


def rule_xlsx_df_pass_fail(df: pd.DataFrame, desc_col: str, val_col: str, skip_on_none=False):
    """
    Processes rows of a DataFrame, checks for specific conditions, and yields results.

    This function iterates over the rows of the input DataFrame and applies a set of
    rules to assess the status of each row, based on the values in the specified
    `desc_col` and `val_col` columns. It detects null values in these columns, and
    depending on the `skip_on_none` argument, either skips processing for those rows
    or treats them as failures. Additionally, it validates the boolean status in the
    `val_col` to determine if a row passes or fails based on its description.

    Args:
        df (pd.DataFrame): Input DataFrame with data rows to be processed.
        desc_col (str): Name of the column in the DataFrame holding descriptive information.
        val_col (str): Name of the column in the DataFrame holding boolean or pass/fail values.
        skip_on_none (bool, optional): If True, rows with null values in the specified columns
            will be skipped rather than treated as failed. Defaults to False.

    Yields:
        TR: An object representing the evaluation result for each row, containing
            the status, whether the row was skipped, and accompanying messages.
    """

    for row in df.values:
        # Make dictionaries because they are easier to look at.
        row_dict = dict(zip(df.columns, row))

        if pd.isnull(row_dict[val_col]):
            if skip_on_none:
                yield TR(status=None, skipped=True,
                         msg=f"Null value detected in column={BM.expected(val_col)}")
            else:
                yield TR(status=False,
                         msg=f"Null value detected in column={BM.expected(val_col)}")
            continue

        if pd.isnull(row_dict[desc_col]):
            if skip_on_none:
                yield TR(status=None, skipped=True,
                         msg=f"Null description detected in column={BM.expected(desc_col)}")
            else:
                yield TR(status=False,
                         msg=f"Null description detected in column={BM.expected(desc_col)}")
            continue

        description = row_dict[desc_col]

        # Very lenient boolean values
        status = str_to_bool(row_dict[val_col])
        if status:
            yield TR(status=True, msg=f"{BM.code(description)}-Passed")
        else:
            yield TR(status=False, msg=f"{BM.code(description)}-Failed")
