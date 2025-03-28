import json
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def convert_to_excel(json_file):
    """将JSON文件转换为Excel格式，并按年份降序排序"""
    try:
        # 读取JSON文件
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 创建输出文件名（基于输入文件名）
        output_file = json_file.rsplit('.', 1)[0] + '_summary.xlsx'
        
        # 准备数据列表
        rows = []
        for article in data['articles']:
            # 处理基因和通路
            if isinstance(article['genes_pathways'], list):
                genes_str = '; '.join(article['genes_pathways'])
            else:
                genes_str = str(article['genes_pathways'])
            
            # 处理研究模型
            if isinstance(article['research_model'], list):
                model_str = '; '.join(article['research_model'])
            else:
                model_str = str(article['research_model'])
            
            # 添加文章信息
            row = {
                '年份': article['publication']['year'],  # 将年份移到前面以便排序
                '标题': article['title'],
                '第一作者': article['authors']['first'],
                '通讯作者': article['authors']['last'],
                '期刊': article['publication']['journal'],
                '结论': article['conclusion'],
                '基因和通路': genes_str,
                '研究模型': model_str
            }
            rows.append(row)
        
        # 创建DataFrame并按年份降序排序
        df = pd.DataFrame(rows)
        df = df.sort_values(by='年份', ascending=False)
        
        # 添加序号列
        df.insert(0, '序号', range(1, len(df) + 1))
        
        # 保存为Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='文献总结')
            
            # 获取工作表
            worksheet = writer.sheets['文献总结']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),  # 数据的最大长度
                    len(str(col))  # 标题的长度
                )
                # 根据内容设置列宽（考虑中英文字符宽度）
                adjusted_width = max_length * 1.5
                # 限制最大列宽
                adjusted_width = min(adjusted_width, 50)
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
            
            # 设置单元格格式
            for row in worksheet.rows:
                for cell in row:
                    # 设置单元格对齐方式
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    # 如果是标题行，设置字体加粗
                    if cell.row == 1:
                        cell.font = Font(bold=True)
            
            # 设置行高
            for row_idx in range(1, worksheet.max_row + 1):
                # 第一行（标题行）设置较小的高度
                if row_idx == 1:
                    worksheet.row_dimensions[row_idx].height = 30
                else:
                    worksheet.row_dimensions[row_idx].height = 60  # 内容行设置更大的高度
        
        print(f"转换完成，结果已保存到 {output_file}")
        print(f"共处理 {len(rows)} 篇文献，已按年份从新到旧排序")
        
    except Exception as e:
        print(f"转换过程中出错: {str(e)}")

def main():
    # 检查是否提供了输入文件参数
    if len(sys.argv) != 2:
        print("使用方法: python 33.py processed_results.json")
        sys.exit(1)
    
    # 获取输入文件名
    input_file = sys.argv[1]
    
    # 处理文件
    convert_to_excel(input_file)

if __name__ == "__main__":
    main()