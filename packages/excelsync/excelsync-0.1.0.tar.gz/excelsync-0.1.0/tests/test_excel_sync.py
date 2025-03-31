"""
Tests for ExcelSync class.
"""

import os
import json
import tempfile
from pathlib import Path

import pytest
import openpyxl

from excelsync import ExcelSync


@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    # Create a temporary Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
    
    # Create a workbook with test data
    wb = openpyxl.Workbook()
    
    # Add data to default sheet
    sheet = wb.active
    sheet.title = "Sheet1"
    
    # Add headers
    sheet["A1"] = "ID"
    sheet["B1"] = "Name"
    sheet["C1"] = "Age"
    sheet["D1"] = "Date"
    
    # Add data
    sheet["A2"] = 1
    sheet["B2"] = "John Doe"
    sheet["C2"] = 30
    sheet["D2"] = "2023-01-01"
    
    sheet["A3"] = 2
    sheet["B3"] = "Jane Smith"
    sheet["C3"] = 28
    sheet["D3"] = "2023-02-15"
    
    # Add a second sheet
    sheet2 = wb.create_sheet(title="Sheet2")
    sheet2["A1"] = "Product"
    sheet2["B1"] = "Price"
    sheet2["A2"] = "Widget"
    sheet2["B2"] = 19.99
    
    # Save the workbook
    wb.save(temp_path)
    
    yield temp_path
    
    # Clean up
    if os.path.exists(temp_path):
        os.unlink(temp_path)


def test_init(sample_excel_file):
    """Test initialization of ExcelSync."""
    excel_sync = ExcelSync(sample_excel_file)
    assert excel_sync.excel_file.exists()
    assert len(excel_sync.workbook.sheetnames) == 2
    assert "Sheet1" in excel_sync.workbook.sheetnames
    assert "Sheet2" in excel_sync.workbook.sheetnames


def test_extract_structure(sample_excel_file):
    """Test extracting structure from Excel file."""
    excel_sync = ExcelSync(sample_excel_file)
    structure = excel_sync.extract_structure()
    
    assert "sheets" in structure
    assert "Sheet1" in structure["sheets"]
    assert "Sheet2" in structure["sheets"]
    
    sheet1 = structure["sheets"]["Sheet1"]
    assert "headers" in sheet1
    assert len(sheet1["headers"]) == 4
    assert sheet1["headers"][1]["name"] == "ID"
    assert sheet1["headers"][2]["name"] == "Name"
    assert sheet1["headers"][3]["name"] == "Age"
    assert sheet1["headers"][4]["name"] == "Date"
    
    # Check data types
    assert sheet1["headers"][1]["data_type"] == "integer"
    assert sheet1["headers"][2]["data_type"] == "string"
    assert sheet1["headers"][3]["data_type"] == "integer"


def test_export_structure(sample_excel_file):
    """Test exporting structure to a file."""
    excel_sync = ExcelSync(sample_excel_file)
    
    with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_structure(temp_path)
    
    # Verify the exported structure
    assert os.path.exists(temp_path)
    with open(temp_path, 'r') as f:
        exported_structure = json.load(f)
    
    assert "sheets" in exported_structure
    assert "Sheet1" in exported_structure["sheets"]
    assert "Sheet2" in exported_structure["sheets"]
    
    # Clean up
    os.unlink(temp_path)


def test_validate_structure(sample_excel_file):
    """Test validating Excel structure."""
    excel_sync = ExcelSync(sample_excel_file)
    structure = excel_sync.extract_structure()
    
    # Structure should be valid against itself
    is_valid, issues = excel_sync.validate_structure(structure)
    assert is_valid
    assert len(issues) == 0
    
    # Modify structure to create an invalid one
    invalid_structure = structure.copy()
    invalid_structure["sheets"]["Sheet1"]["headers"][1]["name"] = "Modified"
    
    is_valid, issues = excel_sync.validate_structure(invalid_structure)
    assert not is_valid
    assert len(issues) > 0


def test_export_to_yaml(sample_excel_file):
    """Test exporting Excel data to YAML."""
    excel_sync = ExcelSync(sample_excel_file)
    
    with tempfile.NamedTemporaryFile(suffix='.yaml', delete=False) as temp:
        temp_path = temp.name
    
    excel_sync.export_to_yaml(temp_path)
    
    # Verify the file exists
    assert os.path.exists(temp_path)
    
    # Clean up
    os.unlink(temp_path) 