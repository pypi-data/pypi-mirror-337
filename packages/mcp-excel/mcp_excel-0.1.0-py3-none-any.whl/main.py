import pandas as pd
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from typing import Dict, Any, Optional, Tuple

mcp = FastMCP(
    "Excel",
    dependencies=["pandas", "openpyxl"]
)

@mcp.tool()
def read_excel(file_path: str, sheet_name: str = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Read an Excel file and return its content as a pandas DataFrame along with its properties.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str, optional): Name or index of the sheet to read. 
                                   If None, reads the first sheet by default.
    
    Returns:
        Tuple[pd.DataFrame, Dict[str, Any]]: A tuple containing:
            - DataFrame containing the Excel sheet data
            - Dictionary containing sheet properties including:
                - data_validation: List of cells with data validation
                - dropdown_lists: List of cells with dropdown lists
                - merged_cells: List of merged cell ranges
                - hidden_rows: List of hidden row numbers
                - hidden_columns: List of hidden column numbers
        
    Raises:
        FileNotFoundError: If the specified file does not exist.
        ValueError: If the specified sheet does not exist in the Excel file.
    """
    try:
        # Read the Excel file for properties
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        
        # Get properties
        properties = {
            "data_validation": [],
            "dropdown_lists": [],
            "merged_cells": list(ws.merged_cells.ranges) if ws.merged_cells else [],
            "hidden_rows": [row for row in range(1, ws.max_row + 1) if ws.row_dimensions[row].hidden],
            "hidden_columns": [col for col in range(1, ws.max_column + 1) if ws.column_dimensions[chr(64 + col)].hidden]
        }
        
        # Get data validation and dropdown lists
        for dv in ws.data_validations.dataValidation:
            # Get the range of cells this validation applies to
            cell_range = dv.sqref  # Use sqref instead of cells
            validation_info = {
                "cell": cell_range,
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allow_blank,
                "show_error": dv.showErrorMessage,
                "show_input": dv.showInputMessage
            }
            properties["data_validation"].append(validation_info)
            
            # Check if it's a dropdown list
            if dv.type == "list":
                # Clean up the formula to get the list options
                options = dv.formula1.strip('"').split(',')
                properties["dropdown_lists"].append({
                    "cell": cell_range,
                    "options": options
                })
        
        # Read the Excel file content
        if sheet_name is None:
            print(f"No specific sheet requested. Reading the first sheet from {file_path}")
            df = pd.read_excel(file_path, engine="openpyxl")
        else:
            print(f"Reading sheet '{sheet_name}' from {file_path}")
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                
        return df, properties
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")
    except ValueError as e:
        raise e
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")

@mcp.tool()
def get_excel_properties(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Get Excel file properties including data validation and dropdown lists.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str, optional): Name of the sheet to analyze. If None, analyzes the first sheet.
    
    Returns:
        Dict[str, Any]: Dictionary containing sheet properties including:
            - data_validation: List of cells with data validation
            - dropdown_lists: List of cells with dropdown lists
            - merged_cells: List of merged cell ranges
            - hidden_rows: List of hidden row numbers
            - hidden_columns: List of hidden column numbers
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")
        
        properties = {
            "data_validation": [],
            "dropdown_lists": [],
            "merged_cells": list(ws.merged_cells.ranges) if ws.merged_cells else [],
            "hidden_rows": [row for row in range(1, ws.max_row + 1) if ws.row_dimensions[row].hidden],
            "hidden_columns": [col for col in range(1, ws.max_column + 1) if ws.column_dimensions[chr(64 + col)].hidden]
        }
        
        # Get data validation and dropdown lists
        for dv in ws.data_validations.dataValidation:
            # Get the range of cells this validation applies to
            cell_range = dv.sqref  # Use sqref instead of cells
            validation_info = {
                "cell": cell_range,
                "type": dv.type,
                "operator": dv.operator,
                "formula1": dv.formula1,
                "formula2": dv.formula2,
                "allow_blank": dv.allow_blank,
                "show_error": dv.showErrorMessage,
                "show_input": dv.showInputMessage
            }
            properties["data_validation"].append(validation_info)
            
            # Check if it's a dropdown list
            if dv.type == "list":
                # Clean up the formula to get the list options
                options = dv.formula1.strip('"').split(',')
                properties["dropdown_lists"].append({
                    "cell": cell_range,
                    "options": options
                })
        
        return properties
        
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")
    except ValueError as e:
        raise e
    except Exception as e:
        raise Exception(f"Error reading Excel properties: {str(e)}")

if __name__ == "__main__":
    mcp.run()