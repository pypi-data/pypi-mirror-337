import tkinter as tk
from tkinter import ttk,TOP,BOTH
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import sys,pandas
import numpy as np
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg,\
    NavigationToolbar2Tk
### Tk App ###

class App:
    '''
    Class to create Graphical User Interface.

    Attributes :

    - root: the interface with a Frame.
    - options: list including the names of optimization variables.
    - df: dataframe including the optimization variables values across iterations.
    - specifications: dataframe presenting the specifications of the optimization problem.

    '''
    def __init__(self, root, options, df, specifications):
        '''
        Initialization of the class App
        :param root: the interface with a Frame
        :param options: list including the names of optimization variables.
        :param df: dataframe including the optimization variables values across iterations.
        :param specifications: dataframe including the specifications of the optimization problem.
        '''
        self.root = root
        self.options = options
        self.selected_options = []  # Keep track of selected options
        self.df = df
        self.specifications = specifications
        self.selected_x = tk.StringVar()
        self.selected_y = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):

        # Create a Treeview widget
        self.treeview = ttk.Treeview(self.root, columns=("Variable", "Min",
                    "Max", "Value", "Type", "Input/Output"), show="headings",
                            height=10)
        self.treeview.pack(padx=10, pady=5)

        # Add column headings
        self.treeview.heading("Variable", text="Variable")
        self.treeview.heading("Min", text="Min")
        self.treeview.heading("Max", text="Max")
        self.treeview.heading("Value", text="Value")
        self.treeview.heading("Type", text="Type")
        self.treeview.heading("Input/Output", text="Input/Output")

        # Define color tags based on type and min/max columns
        type_colors = {
            #"bounds": "black",
            "ineq_cstr": "grey",
            "objective": "red",
            "eq_cstr": "green",
            #"free": "black",
        }

        type_names = {
            "bounds": "constrained",
            "ineq_cstr": "inequality",
            "objective": "objective",
            "eq_cstr": "fixed",
            "free": "free",
        }

        # Add options and additional information
        for option in self.options:
            type_ = self.specifications.Type[option]
            value = self.specifications.Value[option]
            input_output = self.specifications.In_Out[option]

            if type_=='bounds' or type_=='ineq_cstr':
                if isinstance(value[0],list):
                    min_value, max_value=str(np.transpose(np.array(value))[0]),\
                                          str(np.transpose(np.array(value))[1])
                else:
                    min_value, max_value = value
                value_text = ""

            elif type_=='eq_cstr':
                min_value = ""
                max_value = ""
                value_text = str(value)

            else:
                min_value = ""
                max_value = ""
                value_text = ""


            # Add rows and colors
            self.treeview.insert("", "end", values=(option, min_value,
               max_value, value_text, type_names.get(type_), input_output),
                                 tags=(type_,))
            self.treeview.tag_configure(type_, foreground=
            type_colors.get(type_, "black"))

        # Create a scrollbar for the Treeview
        scrollbar = ttk.Scrollbar(self.root, orient="vertical",
                                  command=self.treeview.yview)
        self.treeview.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        # Create a context menu for the Treeview
        self.option_menu = tk.Menu(self.root, tearoff=False)
        self.option_menu.add_command(label="Additional Info",
                                     command=self.show_additional_info)
        self.option_menu.entryconfigure(0, state="disabled")  # Disable the "Additional Info" menu item initially
        self.treeview.bind("<Button-3>", self.show_context_menu)

        self.option_frame = ttk.Frame(self.root)
        self.option_frame.pack(padx=10, pady=10)

        self.x_label = ttk.Label(self.option_frame, text="X:")
        self.x_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.E)

        self.x_combobox = ttk.Combobox(self.option_frame, textvariable=
            self.selected_x, values=self.options)
        self.x_combobox.grid(row=0, column=1, padx=5, pady=5)

        self.y_label = ttk.Label(self.option_frame, text="Y:")
        self.y_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.E)

        self.y_combobox = ttk.Combobox(self.option_frame, textvariable=
            self.selected_y, values=self.options)
        self.y_combobox.grid(row=1, column=1, padx=5, pady=5)

        self.plot_button = ttk.Button(self.option_frame,
                    text="Plot Y / X", command=self.plot_sorted_X)
        self.plot_button.grid(row=3, column=0, columnspan=2, padx=5, pady=10)

        self.export_button = tk.Button(self.root, text="Export",
                                       command=self.export_to_excel)
        self.export_button.pack(pady=5)
        self.canvas=None
        self.toolbar=None
        self.exit_button = tk.Button(self.root, text="Exit",
                                     command=self.exit_program)
        self.exit_button.pack(pady=5)

    def show_context_menu(self, event):
        selection = self.treeview.selection()
        if selection:
            self.selected_options = [self.options[self.treeview.index(item)]
                                     for item in selection]  # Update selected options
            menu_label = f"Additional Info ({len(self.selected_options)} " \
                         f"options selected)"
            self.option_menu.entryconfigure(0, label=menu_label)
            self.option_menu.entryconfigure(0, state="normal")  # Enable the "Additional Info" menu item

            self.option_menu.entryconfigure(1, label="Plot",
                                    command=self.plot_multiple_selected_options)

            self.option_menu.tk_popup(event.x_root, event.y_root)

    def show_additional_info(self):
        for selected_option in self.selected_options:
            print(f"Additional info for {selected_option}")


    def plot_multiple_selected_options(self):
        '''
        Plots multiple selected variables in several graphs.

        :return: /
        '''
        for selected_option in self.selected_options:
            self.plot_one_selected_option(selected_option)

    def export_to_excel(self):
        '''
        Exports the results under Excel format.

        :return: an Excel file including the results of optimization problem.
        '''
        # Create a Workbook object
        workbook = Workbook()

        # Create a new sheet
        sheet = workbook.active

        # Get the columns to export excluding 'IsBestSolution', 'IsSolution', and 'pareto_pts'
        export_columns = [col for col in self.df.columns if col not in
                          ['IsBestSolution', 'IsSolution', 'pareto_pts']]

        # Write the column names to the sheet
        sheet.append(export_columns)

        # Write the row values to the sheet
        for _, row in self.df[export_columns].iterrows():
            values = row.tolist()
            sheet.append(values)

        # Apply conditional text color to the desired columns
        for column in self.specifications.index:
            current_spec = self.specifications.Type[column]
            if current_spec == "bounds":
                spec_values = self.specifications.Value[column]
                min_value, max_value = spec_values[0], spec_values[1]
                font_color = "000000"  # Default black color
                for index, value in self.df[column].items():
                    if value <= min_value:
                        font_color = "0000FF"  # Blue color
                    elif value >= max_value:
                        font_color = "FF0000"  # Red color
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = value
            elif current_spec == "eq_cstr":
                font_color = "808080"  # Grey color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]
            elif current_spec == "objective":
                font_color = "00FF00"  # Green color
                for index in self.df.index:
                    if column in export_columns:
                        cell = sheet.cell(row=index+2,
                                          column=export_columns.index(column)+1)
                        cell.font = Font(color=Color(rgb=font_color))
                        cell.value = self.df[column][index]

        # Save the workbook to a file
        workbook.save("exported_data.xlsx")
        print("Result saved!")

    def plot_one_selected_option(self,selected_option):
        '''
        Plots a selected variable in a graph.

        :param selected_option:
        :return:
        '''

        if isinstance(self.df[selected_option][0],(list,np.ndarray)): #vectorial
            index=[selected_option+str(i) for i in range(len(self.df
                                                [selected_option][0]))]
            df=pandas.DataFrame(np.array(self.df[selected_option].values.
                                         tolist()),columns=index)
            current_spec = self.specifications.Value[selected_option]
            current_type = self.specifications.Type[selected_option]

            if len(index)==2:
                self.fig,self.ax=plt.subplots(2,1)
            else:
                self.fig,self.ax=plt.subplots(2,2)
            self.ax=self.ax.ravel()

            for i in range(len(index)):
                element=index[i]
                if i!=0 and i%4==0:
                    self.display_figures(False)
                    self.fig, self.ax = plt.subplots(2, 2)
                    self.ax = self.ax.ravel()

                self.ax[i%4].plot(self.df['IterationNumber'],df[element],
                             label=element)
                self.ax[i%4].scatter(self.df['IterationNumber'], df[element])

            # Label each dot
                #for iteration, x, y in zip(self.df['IterationNumber'],
                #    self.df['IterationNumber'],df[element]):
                #    self.ax[i%4].annotate(iteration,(x,y),textcoords=
                #       "offset points",xytext=(0, 10), ha='center', va='bottom')

                current_spec_el=current_spec[index.index(element)]
                if current_type=="bounds":
                    self.ax[i%4].axhline(y=current_spec_el[1], color='red',
                                linestyle='--',label=element+'_max')
                    self.ax[i%4].axhline(y=current_spec_el[0],color='blue',
                                linestyle='--',label=element+'_min')

                elif current_type=='ineq_cstr':
                    if current_spec_el[1]!=None:
                        self.ax[i%4].axhline(y=current_spec_el[1],color='black',
                            linestyle='--',label=element + '_max')
                    if current_spec_el[0]!=None:
                        self.ax[i%4].axhline(y=current_spec_el[0], color='grey',
                            linestyle='--',label=element + '_min')

                elif current_type=='eq_cstr':
                    self.ax[i%4].axhline(y=current_spec_el, color='black',
                            linestyle='--',label=element+'_cstr')

                self.ax[i%4].grid()
                self.ax[i%4].set_xlabel("IterationNumber")
                self.ax[i%4].set_title(element)

                # Add the legend
                self.ax[i%4].legend()

                self.ax[i%4].xaxis.set_major_locator(MaxNLocator(integer=True))
                self.ax[i%4].ticklabel_format(useOffset=False, style='plain')

        else: #scalar
            self.fig, self.ax = plt.subplots()
            self.ax.plot(self.df['IterationNumber'],self.df[selected_option],
                    label=selected_option)
            self.ax.scatter(self.df['IterationNumber'],self.df[selected_option])

            # Label each dot
            #for iteration, x, y in zip(self.df['IterationNumber'],
            #            self.df['IterationNumber'], self.df[selected_option]):
            #    self.ax.annotate(iteration, (x, y), textcoords="offset points",
            #                xytext=(0, 10), ha='center', va='bottom')

            current_spec = self.specifications.Value[selected_option]
            current_type = self.specifications.Type[selected_option]

            if current_type=="bounds":
                self.ax.axhline(y=current_spec[1], color='red', linestyle='--',
                           label=selected_option+'_max')
                self.ax.axhline(y=current_spec[0], color='blue', linestyle='--',
                           label=selected_option+'_min')

            elif current_type=='ineq_cstr':
                if current_spec[1]!=None:
                    self.ax.axhline(y=current_spec[1], color='black',
                        linestyle='--',label=selected_option + '_max')
                if current_spec[0]!=None:
                    self.ax.axhline(y=current_spec[0], color='grey',
                        linestyle='--', label=selected_option + '_min')

            elif current_type=='eq_cstr':
                self.ax.axhline(y=current_spec, color='black', linestyle='--',
                           label=selected_option+'_cstr')

            self.ax.grid()
            self.ax.set_xlabel("IterationNumber")
            self.ax.set_title(selected_option)

            # Add the legend
            self.ax.legend()

            self.ax.xaxis.set_major_locator(MaxNLocator(integer=True))
            self.ax.ticklabel_format(useOffset=False, style='plain')

        self.display_figures(False)

    def display_figures(self,block=True):
        '''
        Displays figures of selected variables.

        :param block: bool indicating if figures must be closed before displaying others.
        :return: /
        '''
        #if self.canvas!=None:
        #    self.canvas.get_tk_widget().pack_forget()
        #    self.root.winfo_children()[-1].pack_forget()
        #self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)  # A tk.DrawingArea.
        #self.canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
        #self.toolbar = NavigationToolbar2Tk(self.canvas, self.root)
        #self.canvas.draw_idle()
        plt.show(block=block)

    def plot_sorted_X(self):
        '''
        Plots output optimization variable w.r.t an input optimization variable graph.

        :return: /
        '''

        selected_x = self.selected_x.get()
        selected_y = self.selected_y.get()

        bbox=dict(boxstyle="round", fc="w", alpha=.3)
        arrowprops = dict(arrowstyle="->")

        if isinstance(self.df[selected_x][0],(list,np.ndarray))\
                and not isinstance(self.df[selected_y][0],(list,np.ndarray)):
            # vectorial_x and scalar_y
            index_x=[selected_x+str(i) for i in range(len(self.df
                                                [selected_x][0]))]
            x_values=pandas.DataFrame(np.array(self.df[selected_x].values.
                                         tolist()),columns=index_x)
            y_values=self.df[selected_y]

            if len(index_x) == 2:
                self.fig, self.ax = plt.subplots(2, 1)
            else:
                self.fig, self.ax = plt.subplots(2, 2)
            self.ax = self.ax.ravel()

            for i in range(len(index_x)):
                element_x=index_x[i]
                if i!=0 and i%4==0:
                    self.display_figures(False)
                    self.fig, self.ax = plt.subplots(2, 2)
                    self.ax = self.ax.ravel()
                self.ax[i%4].plot(x_values[element_x], y_values)
                self.ax[i%4].scatter(x_values[element_x], y_values)

                for iteration, x,y in zip(self.df['IterationNumber'],
                                  x_values[element_x],y_values):
                    self.ax[i%4].annotate(iteration,(x, y),textcoords=
                       "offset points",xytext=(0, 10), bbox=bbox,
                                          arrowprops=arrowprops)

                self.ax[i%4].grid()
                self.ax[i%4].set_xlabel(element_x)
                self.ax[i%4].set_ylabel(selected_y)
                self.ax[i%4].set_title(f"Sensitivity of {selected_y}/"
                                       f"{element_x}")

                # Avoid scientific notations
                self.ax[i%4].ticklabel_format(useOffset=False, style='plain')

        elif isinstance(self.df[selected_y][0],(list,np.ndarray))\
                and not isinstance(self.df[selected_x][0],(list,np.ndarray)):
            # vectorial_y and scalar_x
            index_y=[selected_y+str(i) for i in range(len(self.df
                                                [selected_y][0]))]
            y_values=pandas.DataFrame(np.array(self.df[selected_y].values.
                                         tolist()),columns=index_y)
            x_values=self.df[selected_x]

            if len(index_y)==2:
                self.fig,self.ax=plt.subplots(2,1)
            else:
                self.fig,self.ax=plt.subplots(2,2)
            self.ax=self.ax.ravel()

            for i in range(len(index_y)):
                element_y=index_y[i]
                if i!=0 and i%4==0:
                    self.display_figures(False)
                    self.fig, self.ax = plt.subplots(2, 2)
                    self.ax = self.ax.ravel()
                self.ax[i%4].plot(x_values, y_values[element_y])
                self.ax[i%4].scatter(x_values, y_values[element_y])

                for iteration, x,y in zip(self.df['IterationNumber'],
                                  x_values,y_values[element_y]):
                    self.ax[i%4].annotate(iteration,(x, y),textcoords=
                       "offset points",xytext=(0, 10), bbox=bbox,
                                          arrowprops=arrowprops)

                self.ax[i%4].grid()
                self.ax[i%4].set_xlabel(element_y)
                self.ax[i%4].set_ylabel(selected_y)
                self.ax[i%4].set_title(f"Sensitivity of {element_y}/"
                                       f"{selected_x}")

                # Avoid scientific notations
                self.ax[i%4].ticklabel_format(useOffset=False, style='plain')

        elif isinstance(self.df[selected_y][0],(list,np.ndarray))\
                and isinstance(self.df[selected_x][0],(list,np.ndarray)):
            # vectorial_y and vectorial_x
            index_x=[selected_x+str(i) for i in range(len(self.df
                                                [selected_x][0]))]
            index_y=[selected_y+str(i) for i in range(len(self.df
                                                [selected_y][0]))]
            y_values=pandas.DataFrame(np.array(self.df[selected_y].values.
                                         tolist()),columns=index_y)
            x_values=pandas.DataFrame(np.array(self.df[selected_x].values.
                                         tolist()),columns=index_x)

            self.fig, self.ax = plt.subplots(2,2)
            self.ax = self.ax.ravel()

            for i in range(len(index_x)):
                for j in range(len(index_y)):
                    n=i+j+i*(len(index_y)-1)
                    if (n!=0 and n%4==0):
                        self.display_figures(False)
                        self.fig, self.ax = plt.subplots(2, 2)
                        self.ax = self.ax.ravel()
                    element_x,element_y=index_x[i],index_y[j]
                    self.ax[n%4].plot(x_values[element_x], y_values[element_y])
                    self.ax[n%4].scatter(x_values[element_x],
                                         y_values[element_y])

                    for iteration, x,y in zip(self.df['IterationNumber'],
                                  x_values[element_x],y_values[element_y]):
                        self.ax[n%4].annotate(iteration,(x, y),textcoords=
                       "offset points",xytext=(0, 10), bbox=bbox,
                                          arrowprops=arrowprops)

                    self.ax[n%4].grid()
                    self.ax[n%4].set_xlabel(element_x)
                    self.ax[n%4].set_ylabel(element_y)
                    self.ax[n%4].set_title(
                        f"Sensitivity of {element_y}/{element_x}")

                    # Avoid scientific notations
                    self.ax[n%4].ticklabel_format(useOffset=False,style='plain')

        else: #scalar
            sorted_X_df = self.df.sort_values(by=[selected_x])

            # Get the arrays of values for the selected options
            x_values = sorted_X_df[selected_x].values
            y_values = sorted_X_df[selected_y].values

            self.fig, self.ax = plt.subplots()
            self.ax.plot(x_values, y_values)
            self.ax.scatter(x_values, y_values)

            for iteration, x,y in zip(sorted_X_df['IterationNumber'],
                                  x_values,y_values):
                self.ax.annotate(iteration, (x, y), textcoords="offset points",
                        xytext=(0, 10), bbox=bbox,arrowprops=arrowprops)

            self.ax.grid()
            self.ax.set_xlabel(selected_x)
            self.ax.set_ylabel(selected_y)
            self.ax.set_title(f"Sensitivity of {selected_y} / {selected_x}")

            # Avoid scientific notations
            self.ax.ticklabel_format(useOffset=False, style='plain')

        self.display_figures()

    def exit_program(self):
        '''
        Stops the Graphical User Interface.

        :return: /
        '''
        sys.exit()