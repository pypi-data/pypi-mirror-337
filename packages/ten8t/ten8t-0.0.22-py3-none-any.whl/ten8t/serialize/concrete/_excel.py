import pathlib
from typing import Any, List, TextIO

import openpyxl
from openpyxl.styles import Font, PatternFill

from .._base import Ten8tDump
from .._base import Ten8tDumpConfig
from ...ten8t_checker import Ten8tChecker
from ...ten8t_exception import Ten8tException
from ...ten8t_result import Ten8tResult


class Ten8tDumpExcel(Ten8tDump):
    """
    Excel output formatter for Ten8t results.
    Creates Excel workbooks with test results organized in sheets.
    """

    # Base color constants
    COLOR_RED = "FFC7CE"  # Light red
    COLOR_GREEN = "C6EFCE"  # Light green
    COLOR_YELLOW = "FFEB9C"  # Light yellow
    COLOR_BLUE = "DDEBF7"  # Light blue
    COLOR_GRAY = "EDEDED"  # Light gray
    COLOR_PURPLE = "E4D7F5"  # Light purple
    COLOR_ORANGE = "FFE0CC"  # Light orange

    # Semantic color assignments
    COLOR_PASS = COLOR_GREEN
    COLOR_FAIL = COLOR_RED
    COLOR_WARN = COLOR_YELLOW
    COLOR_INFO = COLOR_BLUE
    COLOR_HEADER = COLOR_GRAY

    def __init__(self, config: Ten8tDumpConfig = None):
        if config is None:
            config = Ten8tDumpConfig.excel_default()

        super().__init__(config)

        # Define styles for Excel formatting
        self.header_font = Font(bold=True)
        self.pass_fill = self._make_fill(self.COLOR_PASS)
        self.fail_fill = self._make_fill(self.COLOR_FAIL)
        self.warn_fill = self._make_fill(self.COLOR_WARN)
        self.header_fill = self._make_fill(self.COLOR_GRAY)

        # Get sheet names from config or use defaults
        self.header_sheet_name = self.config.summary_sheet_name
        self.results_sheet_name = self.config.result_sheet_name

    def get_output_file(self, encoding="utf8") -> TextIO:
        # The Excel serializer does not return a IO object because it makes
        # no sense to write to stdout.
        pass

    def dump(self, checker: Ten8tChecker) -> None:
        """
        Custom file writer for Excel output.

        Args:
            checker: Ten8tChecker instance containing results to dump

        Raises:

            Ten8tException: If serialization fails
        """
        output_file = self.config.output_file

        if not output_file:
            raise Ten8tException("No file was specified for xlsx output.")

        try:
            self._dump_implementation(checker, output_file)
        except Exception as e:
            raise Ten8tException(f"Error serializing results: {e}") from e

    @staticmethod
    def _make_fill(color, fill_type="solid"):
        return PatternFill(start_color=color, end_color=color, fill_type=fill_type)

    @staticmethod
    def _fix_column_header(col: str) -> str:
        return col.title().replace("_", "")

    @staticmethod
    def _fix_excel(value: Any) -> str:
        """Make excel compatible value."""
        if isinstance(value, list):
            value = [str(value) for value in value]
            value = ",".join(value)

        if isinstance(value, bool):
            value = 1 if value else 0

        return value

    def _format_header_sheet(self, workbook, header_data: dict, sheet_name='Summary') -> None:
        """
        Populate the header sheet with data from a dictionary.
        Uses a simple approach without manual cell-by-cell writing.

        Args:
            workbook: The Excel workbook object
            header_data: Dictionary containing header metrics (e.g. {"pass_count": 10, "fail_count": 20})
        """
        # Create or get the header sheet (first sheet)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name, 0)

        # Simple row-based population
        row_data = []
        for key, value in header_data.items():
            # Convert snake_case to Title Case for display
            display_name = " ".join(word.capitalize() for word in key.split('_'))
            row_data.append([display_name, self._fix_excel(value)])

        # Write the rows
        for row_idx, (name, value) in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=1, value=name)
            sheet.cell(row=row_idx, column=2, value=value)

        # Set column widths for better readability
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 15

    def _format_results_sheet(self,
                              workbook,
                              results: List[Ten8tResult],
                              sheet_name: str = "Results") -> None:
        """
        Populate the results sheet with data from Ten8tResult objects.
        No manual cell-by-cell writing needed.
        """
        sheet = workbook.create_sheet(sheet_name, 1)

        # Define columns
        columns = self._process_result_columns()

        # Create header row
        header_row = [column.title() for column in columns]

        # Create data rows
        data_rows = []
        for result in results:
            row = []
            for column in columns:
                row.append(self._fix_excel(getattr(result, column, '')))
            data_rows.append(row)

        # Combine header and data rows
        all_rows = [header_row] + data_rows

        # Write all data at once
        for row_idx, row_data in enumerate(all_rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Set column widths (optional)
        sheet.column_dimensions["A"].width = 10  # Status
        sheet.column_dimensions["B"].width = 50  # Message

        # Freeze header row (optional)
        sheet.freeze_panes = "A2"

    def _dump_implementation(self, checker: Ten8tChecker, _: TextIO) -> None:
        """
        Implement the Excel output formatter.

        Args:
            checker: The Ten8tChecker instance containing results
            _: This parameter is not used in this implmentation is not used in this
        """

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Format sheets - using configuration settings
        if self.config.show_summary:
            summary_data = checker.get_header()
            summary_sheet_name = self.config.summary_sheet_name or "Summary"
            self._format_header_sheet(workbook, summary_data, summary_sheet_name)

        if self.config.show_results and checker.results:
            results = checker.results
            result_sheet_name = self.config.result_sheet_name or "Results"
            self._format_results_sheet(workbook, results, result_sheet_name)

        # Determine output file path
        file_path = self.config.output_file or 'result.xlsx'

        # Save workbook
        try:
            file_path = pathlib.Path(file_path)
            if file_path.parent != pathlib.Path('.'):
                file_path.parent.mkdir(parents=True, exist_ok=True)

            workbook.save(file_path)
        except PermissionError as e:
            raise Ten8tException(
                f"Could not write to Excel file: {file_path} - ensure it's not open in another application") from e
        except Exception as e:
            raise Ten8tException(f"Failed to write Excel file: {file_path} - {str(e)}") from e
