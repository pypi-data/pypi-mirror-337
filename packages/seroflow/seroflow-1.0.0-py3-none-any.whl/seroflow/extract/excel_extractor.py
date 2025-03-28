"""
Module: excel_extractor

This module provides concrete implementations for extracting data from Excel files.
It defines two classes:
    - ExcelExtractor: Extracts a DataFrame from a single Excel file.
    - MultiExcelExtractor: Extracts DataFrames from multiple Excel files located in a directory.
These classes extend from the base extractor classes and leverage pandas along with xlrd and
openpyxl to read Excel files.
"""
import xlrd
from openpyxl import load_workbook
import pandas as pd
from ..extract.file_extractor import FileExtractor, MultiFileExtractor

class ExcelExtractor(FileExtractor):
    """
    ExcelExtractor

    A concrete extractor for reading data from an Excel file. This class extends FileExtractor
    and provides methods to read an entire Excel file or a chunk of it, and to add the
    resulting DataFrame to the Pipeline context. It supports both .xls and .xlsx formats
    using the appropriate engines.
    """

    def __init__(self,
                 source,
                 step_name="ExcelExtractor",
                 chunk_size=None,
                 on_error=None,
                 **kwargs):
        """
        ExcelExtractor Class Constructor
        Initializes the ExcelExtractor object.

        Arguments:
            source (str): 
                The source directory where the Excel file is located
            step_name (str): 
                The name of the step
            chunk_size (int): 
                The number of rows to read at a time
            on_error (str): 
                The error handling strategy
            **kwargs: 
                Additional keyword arguments for the read_excel() method
        """
        super().__init__(source=source,
                         step_name=step_name,
                         func = self.func,
                         chunk_size=chunk_size,
                         on_error=on_error,
                         **kwargs)

    def func(self, context):
        """
        Public method: func()
        Reads the Excel file and adds the DataFrame to the context

        Arguments:
            context (Context): 
                Blank context object where the DataFrame will be added

        Returns:
            Context: 
                The context object with the DataFrame added
        """
        context.add_dataframe(self.file_name, self.__read_excel(self.file_path, self.kwargs))
        return context

    def __read_excel(self, file, kwargs):
        """
        Private method: __read_excel()
        Reads the Excel file

        Arguments:
            file (str): 
                The path to the Excel file
            kwargs (dict): 
                Additional keyword arguments for the read_excel() method

        Returns:
            DataFrame: 
                The DataFrame read from the Excel file

        Raises:
            ValueError: 
                If the file format is not supported
        """
        if 'skiprows' in kwargs:
            if kwargs['skiprows'] is None:
                return pd.DataFrame()
        if file.endswith('.xls'):
            return pd.read_excel(file, engine='xlrd', **kwargs)
        if file.endswith('.xlsx'):
            return pd.read_excel(file, engine='openpyxl', **kwargs)
        raise ValueError(f"Unsupported file format: {file}")

    def get_max_row_count(self):
        """
        Public method: get_max_row_count()
        Gets the maximum number of rows in the Excel file

        Returns:
            int: 
                The maximum number of rows in the Excel file
        """
        max_rows = 0
        if self.file_path.endswith('.xlsx'):
            wb = load_workbook(filename=self.file_path, read_only=True)
            ws = wb.active  # use the first (active) sheet
            rows_count = ws.max_row
            wb.close()
        elif self.file_path.endswith('.xls'):
            wb = xlrd.open_workbook(self.file_path, on_demand=True)
            ws = wb.sheet_by_index(0)
            rows_count = ws.nrows
            wb.release_resources()
        else:
            raise ValueError(f"Unsupported file format: {self.file_path}")

        max_rows = max(max_rows, rows_count)
        return max_rows

class MultiExcelExtractor(MultiFileExtractor):
    """
    MultiExcelExtractor

    A concrete extractor for reading data from multiple Excel files located in a directory.
    This class extends MultiFileExtractor and leverages the ExcelExtractor to extract
    DataFrames from each file.
    """
    def __init__(self,
                 source,
                 chunk_size=None,
                 on_error=None,
                 **kwargs):
        """
        MultiExcelExtractor Class Constructor
        Initializes the MultiExcelExtractor object.

        Arguments:
            source (str): 
                The source directory where the Excel files are located
            chunk_size (int): 
                The number of rows to read at a time
            on_error (str): 
                The error handling strategy
            **kwargs: 
                Additional keyword arguments for the ExcelExtractor class
        """
        super().__init__(source=source,
                         step_name="MultiExcelExtractor",
                         type=ExcelExtractor,
                         extension_type='excel',
                         chunk_size=chunk_size,
                         on_error=on_error,
                         **kwargs)
