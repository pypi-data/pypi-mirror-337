# literature_analysis/process.py

import json
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def process_results(input_file: str) -> None:
    """处理分析结果，生成JSON和Excel输出"""
    try:
        # 读取输入文件
        with open(input_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 检查数据结构
        if 'articles' not in data:
            raise ValueError("输入文件格式错误：缺少'articles'字段")

        articles = data['articles']
        
        # 准备Excel数据
        excel_data = []
        for idx, article in enumerate(articles, 1):
            row_data = {
                '序号': idx,
                '年份': article['publication'].get('year', ''),
                '期刊': article['publication'].get('journal', ''),
                '标题': article.get('title', ''),
                '第一作者': article['authors'].get('first', ''),
                '通讯作者': article['authors'].get('last', ''),
                '主要结论': article.get('conclusion', ''),
                '基因和通路': ', '.join(article.get('genes_pathways', [])),
                '研究模型': article.get('research_model', '')
            }
            excel_data.append(row_data)

        # 创建DataFrame
        df = pd.DataFrame(excel_data)

        # 按年份降序排序
        df = df.sort_values(by='年份', ascending=False)
        
        # 重新编号
        df['序号'] = range(1, len(df) + 1)

        # 生成输出文件名
        base_name = input_file.rsplit('.', 1)[0]
        json_output = f"{base_name}_processed.json"
        excel_output = f"{base_name}_summary.xlsx"

        # 保存处理后的JSON
        output_data = {
            "total_count": len(df),
            "articles": df.to_dict('records')
        }
        
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

        # 保存Excel
        writer = pd.ExcelWriter(excel_output, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='文献分析')
        
        # 获取工作表
        worksheet = writer.sheets['文献分析']

        # 设置列宽和格式
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            max_length = min(max_length + 2, 50)  # 限制最大宽度
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = max_length

        # 设置单元格格式
        header_font = Font(bold=True)
        cell_alignment = Alignment(vertical='center', wrap_text=True)

        # 设置标题行格式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = cell_alignment

        # 设置数据行格式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment

        # 设置行高
        worksheet.row_dimensions[1].height = 30  # 标题行高度
        for i in range(2, len(df) + 2):
            worksheet.row_dimensions[i].height = 25  # 内容行高度

        # 保存Excel文件
        writer.close()

        print(f"处理完成！")
        print(f"JSON文件已保存至: {json_output}")
        print(f"Excel文件已保存至: {excel_output}")
        print(f"共处理 {len(df)} 篇文献")

    except Exception as e:
        print(f"处理过程出错: {str(e)}")
        raise

# 如果直接运行此脚本
if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("使用方法: python process.py input_file.json")
        sys.exit(1)

    try:
        process_results(sys.argv[1])
    except Exception as e:
        print(f"执行失败: {str(e)}")
        sys.exit(1)

# 导出函数
__all__ = ['process_results']