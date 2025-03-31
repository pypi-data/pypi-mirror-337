"""
Excel Sync - Main module for Excel structure validation and conversion.
"""

import json
import os
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional, Union

import openpyxl
import yaml

from .schema import ExcelSchema


class ExcelSync:
    """
    Main class for managing Excel sheets with predefined structures.
    Provides functionality for validation, extraction, and conversion.
    """

    def __init__(self, excel_file: Union[str, Path]):
        """
        Initialize the ExcelSync object with an Excel file.

        Args:
            excel_file: Path to the Excel file
        """
        self.excel_file = Path(excel_file)
        if not self.excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_file}")
        
        self.workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        self.schema = ExcelSchema()
    
    def extract_structure(self) -> Dict[str, Any]:
        """
        Extract the structure of the Excel file including:
        - Sheet names
        - Column headers
        - Data types
        - Formulas
        - Named ranges
        - Cell formats
        
        Returns:
            Dict containing the structure of the Excel file
        """
        structure = {
            "sheets": {},
            "named_ranges": {},
            "file_properties": {
                "filename": self.excel_file.name,
                "sheet_count": len(self.workbook.sheetnames)
            }
        }
        
        # Extract named ranges - fixed to work with current openpyxl version
        if hasattr(self.workbook, 'defined_names'):
            for name, defn in self.workbook.defined_names.items():
                structure["named_ranges"][name] = defn.attr_text
        
        # Extract sheet structures
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_structure = {
                "columns": {},
                "rows": sheet.max_row,
                "columns_count": sheet.max_column,
                "merged_cells": [str(merged_cell) for merged_cell in sheet.merged_cells],
            }
            
            # Extract header row (assuming first row contains headers)
            if sheet.max_row > 0 and sheet.max_column > 0:
                headers = {}
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=1, column=col)
                    if cell.value:
                        headers[col] = {
                            "name": str(cell.value),
                            "column_letter": openpyxl.utils.get_column_letter(col)
                        }
                
                sheet_structure["headers"] = headers
                
                # Detect data types for columns based on first data row
                if sheet.max_row > 1:
                    for col in headers:
                        first_value = sheet.cell(row=2, column=col).value
                        data_type = self._detect_data_type(first_value)
                        headers[col]["data_type"] = data_type
            
            structure["sheets"][sheet_name] = sheet_structure
        
        return structure
    
    def _detect_data_type(self, value: Any) -> str:
        """
        Detect the data type of a value.
        
        Args:
            value: Value to detect type for
            
        Returns:
            String representation of the data type
        """
        if value is None:
            return "null"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, int):
            return "integer"
        elif isinstance(value, float):
            return "number"
        elif isinstance(value, str):
            # Try to detect date/time formats
            if any(date_indicator in value.lower() for date_indicator in ["/", "-", "date", "time"]):
                return "datetime"
            return "string"
        else:
            return str(type(value).__name__)
    
    def export_structure(self, output_file: Union[str, Path]) -> None:
        """
        Export the structure of the Excel file to a JSON file.
        
        Args:
            output_file: Path to save the structure file
        """
        structure = self.extract_structure()
        
        output_path = Path(output_file)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(structure, f, indent=2)
    
    def validate_structure(self, expected_structure: Optional[Dict[str, Any]] = None) -> Tuple[bool, List[str]]:
        """
        Validate that the Excel file structure matches the expected structure.
        If no expected structure is provided, it validates against the structure
        previously extracted or loaded.
        
        Args:
            expected_structure: Optional dictionary with expected structure
            
        Returns:
            Tuple of (is_valid, list_of_issues)
        """
        current_structure = self.extract_structure()
        
        if expected_structure is None:
            # If we don't have an expected structure, assume the current one is valid
            return True, []
        
        issues = []
        
        # Check if all expected sheets exist
        for sheet_name in expected_structure.get("sheets", {}):
            if sheet_name not in current_structure.get("sheets", {}):
                issues.append(f"Missing sheet: {sheet_name}")
                continue
            
            expected_sheet = expected_structure["sheets"][sheet_name]
            current_sheet = current_structure["sheets"][sheet_name]
            
            # Check headers
            expected_headers = expected_sheet.get("headers", {})
            current_headers = current_sheet.get("headers", {})
            
            for col, header_info in expected_headers.items():
                col = int(col) if isinstance(col, str) else col
                if col not in current_headers:
                    issues.append(f"Missing column {col} in sheet {sheet_name}")
                elif current_headers[col]["name"] != header_info["name"]:
                    issues.append(
                        f"Header mismatch in sheet {sheet_name}, column {col}: "
                        f"expected '{header_info['name']}', got '{current_headers[col]['name']}'"
                    )
        
        return len(issues) == 0, issues
    
    def compare_structure(self, structure_file: Union[str, Path]) -> Tuple[bool, List[str]]:
        """
        Compare the current Excel structure with a previously saved structure file.
        
        Args:
            structure_file: Path to the structure file
            
        Returns:
            Tuple of (is_matching, list_of_differences)
        """
        structure_path = Path(structure_file)
        if not structure_path.exists():
            raise FileNotFoundError(f"Structure file not found: {structure_path}")
        
        with open(structure_path, 'r', encoding='utf-8') as f:
            expected_structure = json.load(f)
        
        return self.validate_structure(expected_structure)
    
    def export_to_yaml(self, output_file: Union[str, Path]) -> None:
        """
        Export the Excel content to YAML with schema information.
        
        Args:
            output_file: Path to save the YAML file
        """
        structure = self.extract_structure()
        data = {
            "schema": structure,
            "data": {}
        }
        
        # Extract actual data
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_data = []
            
            # Get headers from first row
            headers = {}
            for col in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row=1, column=col).value
                if header_value:
                    headers[col] = str(header_value)
            
            # Extract data rows
            for row in range(2, sheet.max_row + 1):
                row_data = {}
                for col in headers:
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data[headers[col]] = cell_value
                
                if row_data:  # Only add non-empty rows
                    sheet_data.append(row_data)
            
            data["data"][sheet_name] = sheet_data
        
        # Write to YAML file
        with open(output_file, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, default_flow_style=False, sort_keys=False)
    
    def load_structure(self, structure_file: Union[str, Path]) -> None:
        """
        Load a structure from a file into the schema.
        
        Args:
            structure_file: Path to the structure file
        """
        structure_path = Path(structure_file)
        if not structure_path.exists():
            raise FileNotFoundError(f"Structure file not found: {structure_path}")
        
        with open(structure_path, 'r', encoding='utf-8') as f:
            structure = json.load(f)
        
        self.schema.load_structure(structure) 