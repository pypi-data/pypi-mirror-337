# coding=utf-8

from typing import Any
from typing import List
from typing import Optional
from typing import Tuple

import openpyxl
from xkits_file.safefile import SafeFile

from xkits_sheet.table import Form


class XLSX():
    """Read or write .xlsx file
    """

    def __init__(self, filename: str, read_only: bool = True):
        self.__file: str = filename
        SafeFile.restore(path=filename)
        self.__book: openpyxl.Workbook = openpyxl.load_workbook(
            filename=filename, read_only=read_only)

    @property
    def file(self) -> str:
        return self.__file

    @property
    def book(self) -> openpyxl.Workbook:
        return self.__book

    def load_sheet(self, sheet_name: Optional[str] = None) -> Form[str, Any]:
        def get_default_sheet_name() -> str:
            if isinstance(sheet_name, str):
                return sheet_name
            active = self.book.active
            return self.book.sheetnames[0] if active is None else active.title

        sheet = self.book[get_default_sheet_name()]
        first = list(sheet.iter_rows(max_row=1))[0]
        cells: List[str] = [c.value for c in first if isinstance(c.value, str)]
        table: Form[str, Any] = Form(name=sheet.title, header=cells)
        for _row in sheet.iter_rows(min_row=2):
            table.append([cell.value for cell in _row])
        return table

    def load_sheets(self) -> Tuple[Form[str, str], ...]:
        return tuple(self.load_sheet(name) for name in self.book.sheetnames)
